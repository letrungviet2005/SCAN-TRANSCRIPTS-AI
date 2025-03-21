import os
import uuid
from openpyxl import Workbook
from openpyxl.styles import PatternFill

def export_to_excel(data_list, output_dir="./results_excel"):
    if not os.path.exists(output_dir):
        os.makedirs(output_dir)

    wb = Workbook()
    ws = wb.active
    ws.title = "Student Scores"

    ws.append(["ID", "Scores"])

    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

    flattened_data = []
    for sublist in data_list:
        if isinstance(sublist, list):
            flattened_data.extend(sublist)
        else:
            flattened_data.append(sublist)

    for row in flattened_data:
        if not isinstance(row, list) or len(row) < 2:
            print("Skipping invalid row:", row)  # Debugging
            continue

        student_id = row[0].get("text", "Unknown")
        score = row[1].get("text", "Unknown")
        confidence = row[1].get("confidence", 1.0)
        is_match = row[1].get("is_match", False)

        excel_row = [student_id, score]
        ws.append(excel_row)

        if is_match == False:
            ws.cell(row=ws.max_row, column=2).fill = yellow_fill

    file_path = os.path.join(output_dir, f"student_scores_{uuid.uuid4().hex}.xlsx")
    wb.save(file_path)
    return file_path